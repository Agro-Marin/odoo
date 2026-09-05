"""Text out of the four office containers, for anything holding one.

These parsers were methods on `ir.attachment` and reachable only through
`_index()`, so a document layer that had every other format could not read a
Word document. They are plain functions here, registered as readers of the
shared registry, and `ir.attachment` calls the same functions with its own
zip-entry bound.

The bound is a parameter rather than a constant either side owns. The indexer
passes `_INDEX_MAX_BYTES`, which is also its read size and its stored-content
limit; a reader reached through the registry has no such budget to inherit and
takes `MAX_ENTRY_BYTES`. Naming one number for both would tie a zip-bomb guard
to how much text a column happens to hold.
"""

from __future__ import annotations

import io
import logging
import re
import warnings
import xml.dom
import zipfile

from defusedxml.minidom import parseString as defused_parse_string
from lxml import etree

from odoo.libs.documents import (
    TEXT,
    BaseReader,
    mimetype_for,
    mimetypes_for,
    register_reader,
)

_logger = logging.getLogger(__name__)

DOCX = mimetype_for("docx")
PPTX = mimetype_for("pptx")
XLSX = mimetype_for("xlsx")
OPENDOCUMENT = mimetypes_for("odt", "ods", "odp", "odg")

# What one entry of a zip-based container may inflate to. A small, well-formed
# .docx can declare a huge uncompressed size for its inner XML and force a full
# in-memory inflate of attacker-controlled content on every read.
MAX_ENTRY_BYTES = 4 * 1024 * 1024

MAX_COLUMN_REPEAT = 100
MAX_ROW_REPEAT = 50

_ODF_NAMESPACES = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "manifest": "urn:oasis:names:tc:opendocument:xmlns:manifest:1.0",
}


def text_to_string(element):
    buff = ""
    for node in element.childNodes:
        if node.nodeType == xml.dom.Node.TEXT_NODE:
            buff += node.nodeValue
        elif node.nodeType == xml.dom.Node.ELEMENT_NODE:
            buff += text_to_string(node)
    return buff


def clean_text_content(buf):
    """Drop NULs and CRs, tabs to spaces, collapse whitespace."""
    if not buf:
        return buf
    buf = buf.translate(
        {
            ord("\x00"): None,
            ord("\r"): None,
            ord("\t"): ord(" "),
        }
    )

    def _compact_whitespace(match):
        chunk = match.group(0)
        newline_count = chunk.count("\n")
        if newline_count == 0:
            return " "
        return "\n\n" if newline_count > 1 else "\n"

    buf = re.sub(r"\s{2,}", _compact_whitespace, buf)
    return buf.strip()


def csv_escape(value):
    if value is None:
        return ""
    value = str(value)
    if "," in value or '"' in value or "\n" in value or "\r" in value:
        return '"' + value.replace('"', '""') + '"'
    return value


def read_zip_entry(zf, info, max_entry_bytes=MAX_ENTRY_BYTES):
    if info.file_size > max_entry_bytes:
        _logger.info(
            "attachment_indexation: skipping oversized zip entry %r (%d bytes)",
            info.filename,
            info.file_size,
        )
        return None
    return zf.read(info)


def read_docx(data, max_entry_bytes=MAX_ENTRY_BYTES):
    buf = ""
    f = io.BytesIO(data)
    if zipfile.is_zipfile(f):
        try:
            zf = zipfile.ZipFile(f)
            raw = read_zip_entry(zf, zf.getinfo("word/document.xml"), max_entry_bytes)
            if raw is None:
                return buf
            content = defused_parse_string(raw)
            for element in content.getElementsByTagName("w:p"):
                buf += text_to_string(element) + "\n"
        except Exception:
            _logger.debug(
                "attachment_indexation: failed to index docx content", exc_info=True
            )
    return buf


def read_pptx(data, max_entry_bytes=MAX_ENTRY_BYTES):
    buf = ""
    f = io.BytesIO(data)
    if zipfile.is_zipfile(f):
        try:
            zf = zipfile.ZipFile(f)
            zf_filelist = [x for x in zf.namelist() if x.startswith("ppt/slides/slide")]
            for name in zf_filelist:
                raw = read_zip_entry(zf, zf.getinfo(name), max_entry_bytes)
                if raw is None:
                    continue
                content = defused_parse_string(raw)
                for element in content.getElementsByTagName("a:t"):
                    buf += text_to_string(element) + "\n"
        except Exception:
            _logger.debug(
                "attachment_indexation: failed to index pptx content", exc_info=True
            )
    return buf


def read_xlsx(data, max_entry_bytes=MAX_ENTRY_BYTES):
    try:
        from openpyxl import load_workbook

        logging.getLogger("openpyxl").setLevel(logging.CRITICAL)
    except ImportError:
        _logger.info("openpyxl is not installed.")
        return ""

    f = io.BytesIO(data)
    if zipfile.is_zipfile(f):
        # Only the parts load_workbook actually decompresses to resolve cells
        # (sheets, shared strings, styles, the workbook manifest itself) can
        # zip-bomb it. xl/media and xl/embeddings hold images and OLE objects
        # respectively, which a legitimate spreadsheet may carry at any size
        # without load_workbook ever inflating them.
        oversized = any(
            info.file_size > max_entry_bytes
            for info in zipfile.ZipFile(f).infolist()
            if info.filename.startswith("xl/")
            and not info.filename.startswith(("xl/media/", "xl/embeddings/"))
        )
        f.seek(0)
        if oversized:
            _logger.info(
                "attachment_indexation: skipping oversized xlsx zip entry "
                "(zip-bomb guard)"
            )
            return ""

    all_sheets = []
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = load_workbook(f, data_only=True, read_only=True)
            for sheet in workbook.worksheets:
                sheet_name_escaped = csv_escape(sheet.title)
                sheet_rows = []
                for row in sheet.iter_rows(values_only=True):
                    if not any(row):
                        continue
                    row_cells = [sheet_name_escaped] + [
                        csv_escape(str(cell) if cell is not None else "")
                        for cell in row
                    ]
                    sheet_rows.append(",".join(row_cells))
                sheet_data = "\n".join(sheet_rows)
                if sheet_data:
                    all_sheets.append(sheet_data)
    except Exception:
        _logger.debug(
            "attachment_indexation: failed to index xlsx content", exc_info=True
        )

    return clean_text_content("\n\n".join(all_sheets))


def _extract_row(row):
    cells = []
    for cell in row.xpath(
        ".//table:table-cell | .//table:covered-table-cell",
        namespaces=_ODF_NAMESPACES,
    ):
        repeat = cell.get(f"{{{_ODF_NAMESPACES['table']}}}number-columns-repeated")
        repeat_count = (
            min(int(repeat), MAX_COLUMN_REPEAT) if repeat and repeat.isdigit() else 1
        )
        text_parts = cell.xpath(".//text:p//text()", namespaces=_ODF_NAMESPACES)
        cell_text = " ".join(t.strip() for t in text_parts if t.strip())
        cells.extend([cell_text] * repeat_count)
    return cells


def _extract_spreadsheet(content):
    sheets_csv = []
    for table in content.xpath(".//table:table", namespaces=_ODF_NAMESPACES):
        table_rows = []
        table_name = table.get(f"{{{_ODF_NAMESPACES['table']}}}name")
        if not table_name:
            table_name = f"Sheet{len(sheets_csv) + 1}"
        table_name_escaped = csv_escape(table_name)
        for row in table.xpath(".//table:table-row", namespaces=_ODF_NAMESPACES):
            row_repeat = row.get(f"{{{_ODF_NAMESPACES['table']}}}number-rows-repeated")
            row_repeat_count = (
                min(int(row_repeat), MAX_ROW_REPEAT)
                if row_repeat and row_repeat.isdigit()
                else 1
            )
            cells = _extract_row(row)
            if not any(cells):
                continue
            while cells and not cells[-1]:
                cells.pop()
            row_str = ",".join([table_name_escaped] + list(map(csv_escape, cells)))
            if row_str.replace(",", "").strip():
                table_rows.extend([row_str] * row_repeat_count)
        if table_rows:
            sheets_csv.append("\n".join(table_rows))
    return sheets_csv


def _extract_text(content):
    lines = []
    for element in content.xpath(
        ".//text:p | .//text:h | .//text:list-item", namespaces=_ODF_NAMESPACES
    ):
        text = "".join(element.xpath(".//text()", namespaces=_ODF_NAMESPACES)).strip()
        if text:
            lines.append(text)
    return lines


def read_opendoc(data, max_entry_bytes=MAX_ENTRY_BYTES):
    f = io.BytesIO(data)
    buf = []
    if zipfile.is_zipfile(f):
        try:
            zf = zipfile.ZipFile(f)
            raw = read_zip_entry(zf, zf.getinfo("content.xml"), max_entry_bytes)
            if raw is None:
                return clean_text_content("")
            # Explicit hardened parser (defense-in-depth): don't rely solely on
            # the process-wide lxml default to guard against XXE and entity
            # expansion.
            parser = etree.XMLParser(resolve_entities=False, no_network=True)
            content = etree.fromstring(raw, parser=parser)
            mime_type = zf.read("mimetype").decode("utf-8").strip()
            if mime_type and "spreadsheet" in mime_type:
                buf.extend(_extract_spreadsheet(content))
            else:
                buf.extend(_extract_text(content))
        except Exception:
            _logger.debug(
                "attachment_indexation: failed to index opendoc content", exc_info=True
            )

    return clean_text_content("\n\n".join(buf))


class _OfficeText(BaseReader):
    """One office container's text, for whoever holds the bytes."""

    yields = (TEXT,)

    def __init__(self, name, mimetypes, read):
        self.name = name
        self.mimetypes = frozenset(mimetypes)
        self._read = read

    def read(self, document):
        return self._read(
            document.data,
            document.options.get("max_zip_entry_bytes", MAX_ENTRY_BYTES),
        )


register_reader(_OfficeText("docx_text", {DOCX}, read_docx))
register_reader(_OfficeText("pptx_text", {PPTX}, read_pptx))
register_reader(_OfficeText("xlsx_text", {XLSX}, read_xlsx))
register_reader(_OfficeText("opendocument_text", OPENDOCUMENT, read_opendoc))
