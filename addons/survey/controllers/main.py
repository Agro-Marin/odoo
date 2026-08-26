import contextlib
import csv
import io
import json
import logging
from collections import defaultdict
from datetime import timedelta
from typing import Any
from urllib.parse import quote

import werkzeug
from dateutil.relativedelta import relativedelta
from werkzeug.wrappers import Response

from odoo import _, fields, http
from odoo.exceptions import AccessError, UserError
from odoo.fields import Domain
from odoo.http import content_disposition, request
from odoo.tools import format_date, format_datetime, is_html_empty
from odoo.tools.urls import keep_query

_logger = logging.getLogger(__name__)


class Survey(http.Controller):
    MAX_UPLOADS_PER_ANSWER = 20

    def _fetch_from_access_token(
        self, survey_token: str, answer_token: str | bool
    ) -> tuple[Any, Any]:
        SurveySudo, UserInputSudo = (
            request.env["survey.survey"].sudo(),
            request.env["survey.user_input"].sudo(),
        )
        if not survey_token:
            return SurveySudo, UserInputSudo
        if answer_token:
            answer_sudo = UserInputSudo.search(
                Domain(
                    "survey_id",
                    "any",
                    Domain("access_token", "=", survey_token)
                    & Domain("active", "in", (True, False)),
                )
                & Domain("access_token", "=", answer_token),
                limit=1,
            )
            if answer_sudo:
                return answer_sudo.survey_id, answer_sudo

        return SurveySudo.with_context(active_test=False).search(
            [("access_token", "=", survey_token)]
        ), UserInputSudo

    def _check_validity(
        self,
        survey_sudo: Any,
        answer_sudo: Any,
        answer_token: str | None,
        ensure_token: bool = True,
        check_partner: bool = True,
    ) -> str | bool:
        if not survey_sudo:
            return "survey_wrong"

        if answer_token and not answer_sudo:
            return "token_wrong"

        if not answer_sudo and ensure_token:
            return "token_required"
        if not answer_sudo and survey_sudo.access_mode == "token":
            return "token_required"

        if survey_sudo.users_login_required and request.env.user._is_public():
            return "survey_auth"

        if not survey_sudo.active and (not answer_sudo or not answer_sudo.test_entry):
            return "survey_closed"

        if (
            not survey_sudo.page_ids
            and survey_sudo.questions_layout_effective == "page_per_section"
        ) or not survey_sudo.question_ids:
            return "survey_void"

        if (
            answer_sudo
            and answer_sudo.deadline
            and answer_sudo.deadline < fields.Datetime.now()
        ):
            return "answer_deadline"

        if answer_sudo and check_partner:
            if (
                request.env.user._is_public()
                and answer_sudo.partner_id
                and not answer_token
            ):
                return "answer_wrong_user"
            if (
                not request.env.user._is_public()
                and answer_sudo.partner_id != request.env.user.partner_id
            ):
                return "answer_wrong_user"

        return True

    def _get_access_data(
        self,
        survey_token: str,
        answer_token: str | None,
        ensure_token: bool = True,
        check_partner: bool = True,
    ) -> dict[str, Any]:
        survey_sudo, answer_sudo = self._fetch_from_access_token(
            survey_token, answer_token
        )
        has_survey_access, can_answer = False, False

        validity_code = self._check_validity(
            survey_sudo,
            answer_sudo,
            answer_token,
            ensure_token=ensure_token,
            check_partner=check_partner,
        )
        if validity_code != "survey_wrong":
            has_survey_access = survey_sudo.with_user(request.env.user).has_access(
                "read"
            )
            can_answer = bool(answer_sudo)
            if not can_answer:
                can_answer = survey_sudo.access_mode == "public"

        return {
            "survey_sudo": survey_sudo,
            "answer_sudo": answer_sudo,
            "has_survey_access": has_survey_access,
            "can_answer": can_answer,
            "validity_code": validity_code,
        }

    def _redirect_with_error(
        self, access_data: dict[str, Any], error_key: str
    ) -> Response:
        survey_sudo = access_data["survey_sudo"]
        answer_sudo = access_data["answer_sudo"]

        if error_key == "survey_void" and access_data["can_answer"]:
            return request.render(
                "survey.survey_void_content",
                {"survey": survey_sudo, "answer": answer_sudo},
            )
        elif error_key == "survey_closed" and access_data["can_answer"]:
            return request.render(
                "survey.survey_closed_expired", {"survey": survey_sudo}
            )
        elif error_key == "survey_auth":
            if not answer_sudo:
                redirect_url = (
                    f"/web/login?redirect=/survey/start/{survey_sudo.access_token}"
                )
            elif answer_sudo.access_token:
                if answer_sudo.partner_id and (
                    answer_sudo.partner_id.user_ids or survey_sudo.users_can_signup
                ):
                    if answer_sudo.partner_id.user_ids:
                        answer_sudo.partner_id.signup_cancel()
                    else:
                        answer_sudo.partner_id.signup_prepare()
                    redirect_url = answer_sudo.partner_id._get_signup_url_for_action(
                        url=f"/survey/start/{survey_sudo.access_token}?answer_token={answer_sudo.access_token}"
                    )[answer_sudo.partner_id.id]
                else:
                    survey_url = f"/survey/start/{survey_sudo.access_token}?answer_token={answer_sudo.access_token}"
                    redirect_url = f"/web/login?redirect={quote(survey_url, safe='')}"
            return request.render(
                "survey.survey_auth_required",
                {"survey": survey_sudo, "redirect_url": redirect_url},
            )
        elif error_key == "answer_deadline" and answer_sudo.access_token:
            return request.render(
                "survey.survey_closed_expired", {"survey": survey_sudo}
            )
        elif error_key in ["answer_wrong_user", "token_wrong"]:
            return request.render("survey.survey_access_error", {"survey": survey_sudo})

        return request.redirect("/")

    @http.route(
        "/survey/test/<string:survey_token>", type="http", auth="user", website=True
    )
    def survey_test(self, survey_token: str, **kwargs: Any) -> Response:
        survey_sudo, _dummy = self._fetch_from_access_token(survey_token, False)
        try:
            answer_sudo = survey_sudo._create_answer(
                user=request.env.user, test_entry=True
            )
        except AccessError, UserError:
            return request.redirect("/")
        return request.redirect(
            f"/survey/start/{survey_sudo.access_token}?{keep_query('*', answer_token=answer_sudo.access_token)}"
        )

    @http.route(
        "/survey/retry/<string:survey_token>/<string:answer_token>",
        type="http",
        auth="public",
        website=True,
    )
    def survey_retry(
        self, survey_token: str, answer_token: str, **post: Any
    ) -> Response:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=True
        )
        if access_data["validity_code"] is not True:
            return self._redirect_with_error(access_data, access_data["validity_code"])

        survey_sudo, answer_sudo = (
            access_data["survey_sudo"],
            access_data["answer_sudo"],
        )
        if not answer_sudo:
            return request.redirect("/")

        try:
            retry_answer_sudo = survey_sudo._create_answer(
                user=request.env.user,
                partner=answer_sudo.partner_id,
                email=answer_sudo.email,
                invite_token=answer_sudo.invite_token,
                test_entry=answer_sudo.test_entry,
                **self._prepare_retry_additional_values(answer_sudo),
            )
        except AccessError, UserError:
            return request.redirect("/")
        return request.redirect(
            f"/survey/start/{survey_sudo.access_token}?{keep_query('*', answer_token=retry_answer_sudo.access_token)}"
        )

    def _prepare_retry_additional_values(self, answer: Any) -> dict[str, Any]:
        return {
            "deadline": answer.deadline,
            "nickname": answer.nickname,
        }

    def _prepare_survey_finished_values(
        self, survey: Any, answer: Any, token: str | bool = False
    ) -> dict[str, Any]:
        values = {"survey": survey, "answer": answer}
        if token:
            values["token"] = token
        return values

    @http.route(
        "/survey/start/<string:survey_token>", type="http", auth="public", website=True
    )
    def survey_start(
        self,
        survey_token: str,
        answer_token: str | None = None,
        email: str | bool = False,
        **post: Any,
    ) -> Response:
        answer_from_cookie = False
        if not answer_token:
            answer_token = request.cookies.get(f"survey_{survey_token}")
            answer_from_cookie = bool(answer_token)

        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=False
        )

        if answer_from_cookie and access_data["validity_code"] in (
            "answer_wrong_user",
            "token_wrong",
        ):
            access_data = self._get_access_data(survey_token, None, ensure_token=False)

        if access_data["validity_code"] is not True:
            return self._redirect_with_error(access_data, access_data["validity_code"])

        survey_sudo, answer_sudo = (
            access_data["survey_sudo"],
            access_data["answer_sudo"],
        )
        if not answer_sudo:
            try:
                answer_sudo = survey_sudo._create_answer(
                    user=request.env.user, email=email
                )
            except UserError:
                answer_sudo = False

        if not answer_sudo:
            try:
                survey_sudo.with_user(request.env.user).check_access("read")
            except AccessError:
                return request.redirect("/")
            else:
                return request.render("survey.survey_403_page", {"survey": survey_sudo})

        lang = self._get_lang_with_fallback(answer_sudo.sudo(False))
        url_from = f"/survey/{survey_sudo.access_token}/{answer_sudo.access_token}"
        return request.redirect(self.env["ir.http"]._url_for(url_from, lang.code))

    def _prepare_survey_data(
        self, survey_sudo: Any, answer_sudo: Any, **post: Any
    ) -> dict[str, Any]:
        data = self._prepare_survey_base_data(survey_sudo, answer_sudo)
        (
            triggering_answers_by_question,
            triggered_questions_by_answer,
            selected_answers,
        ) = answer_sudo._get_conditional_values()
        if survey_sudo.questions_layout_effective != "page_per_question":
            data.update(
                {
                    "triggering_answers_by_question": {
                        question.id: triggering_answers.ids
                        for question, triggering_answers in triggering_answers_by_question.items()
                        if triggering_answers
                    },
                    "triggered_questions_by_answer": {
                        answer.id: triggered_questions.ids
                        for answer, triggered_questions in triggered_questions_by_answer.items()
                    },
                    "selected_answers": selected_answers.ids,
                }
            )

        page_or_question_key = (
            "question"
            if survey_sudo.questions_layout_effective == "page_per_question"
            else "page"
        )

        if "previous_page_id" in post:
            return self._prepare_survey_back_navigation_data(
                data, survey_sudo, answer_sudo, page_or_question_key, post
            )

        if answer_sudo.state == "in_progress":
            self._prepare_survey_in_progress_data(
                data,
                survey_sudo,
                answer_sudo,
                page_or_question_key,
                triggered_questions_by_answer,
                post,
            )
        elif answer_sudo.state == "done" or answer_sudo.survey_time_limit_reached:
            return self._prepare_survey_finished_values(survey_sudo, answer_sudo)

        return data

    def _prepare_survey_base_data(
        self, survey_sudo: Any, answer_sudo: Any
    ) -> dict[str, Any]:
        data = {
            "is_html_empty": is_html_empty,
            "survey": survey_sudo,
            "answer": answer_sudo,
            "skipped_questions": answer_sudo._get_skipped_questions(),
            "breadcrumb_pages": [
                {"id": page.id, "title": page.title} for page in survey_sudo.page_ids
            ],
            "format_datetime": lambda dt: format_datetime(
                request.env, dt, dt_format=False
            ),
            "format_date": lambda date: format_date(request.env, date),
        }
        if answer_sudo.state == "new":
            supported_lang_codes = survey_sudo._get_supported_lang_codes()
            data["languages"] = [
                (lang_code, self.env["res.lang"]._get_data(code=lang_code)["name"])
                for lang_code in supported_lang_codes
            ]
            data["lang_code"] = self._get_lang_with_fallback(
                answer_sudo.sudo(False)
            ).code

        if (
            not answer_sudo.is_session_answer
            and survey_sudo.is_time_limited
            and answer_sudo.start_datetime
        ):
            data.update(
                {
                    "server_time": fields.Datetime.now(),
                    "timer_start": answer_sudo.start_datetime.isoformat(),
                    "time_limit_minutes": survey_sudo.time_limit,
                }
            )
        return data

    def _prepare_survey_back_navigation_data(
        self,
        data: dict[str, Any],
        survey_sudo: Any,
        answer_sudo: Any,
        page_or_question_key: str,
        post: dict[str, Any],
    ) -> dict[str, Any]:
        try:
            previous_page_or_question_id = int(post["previous_page_id"])
        except ValueError, TypeError:
            return data
        if previous_page_or_question_id not in survey_sudo.question_and_page_ids.ids:
            return data
        new_previous_id = survey_sudo._get_next_page_or_question(
            answer_sudo, previous_page_or_question_id, go_back=True
        ).id
        page_or_question = (
            request.env["survey.question"].sudo().browse(previous_page_or_question_id)
        )
        data.update(
            {
                page_or_question_key: page_or_question,
                "previous_page_id": new_previous_id,
                "has_answered": answer_sudo.user_input_line_ids.filtered(
                    lambda line: line.question_id.id == new_previous_id
                ),
                "can_go_back": survey_sudo._can_go_back(answer_sudo, page_or_question),
            }
        )
        return data

    def _prepare_survey_in_progress_data(
        self,
        data: dict[str, Any],
        survey_sudo: Any,
        answer_sudo: Any,
        page_or_question_key: str,
        triggered_questions_by_answer: Any,
        post: dict[str, Any],
    ) -> None:
        next_page_or_question = self._resolve_next_page_or_question(
            survey_sudo, answer_sudo, post
        )

        if next_page_or_question:
            if answer_sudo.survey_first_submitted:
                survey_last = answer_sudo._is_last_skipped_page_or_question(
                    next_page_or_question
                )
            else:
                survey_last = survey_sudo._is_last_page_or_question(
                    answer_sudo, next_page_or_question
                )
            data["survey_last"] = survey_last

            if (
                not answer_sudo.survey_first_submitted
                and survey_last
                and survey_sudo.questions_layout_effective != "one_page"
            ):
                data["survey_last_triggering_answers"] = (
                    self._get_last_page_triggering_answers(
                        survey_sudo,
                        answer_sudo,
                        next_page_or_question,
                        triggered_questions_by_answer,
                    )
                )

        if answer_sudo.is_session_answer and next_page_or_question.is_time_limited:
            data.update(
                {
                    "timer_start": survey_sudo.session_question_start_time.isoformat(),
                    "time_limit_minutes": next_page_or_question.time_limit / 60,
                }
            )

        data.update(
            {
                page_or_question_key: next_page_or_question,
                "has_answered": answer_sudo.user_input_line_ids.filtered(
                    lambda line: line.question_id == next_page_or_question
                ),
                "can_go_back": survey_sudo._can_go_back(
                    answer_sudo, next_page_or_question
                ),
            }
        )
        if survey_sudo.questions_layout_effective != "one_page":
            data["previous_page_id"] = survey_sudo._get_next_page_or_question(
                answer_sudo, next_page_or_question.id, go_back=True
            ).id

    def _resolve_next_page_or_question(
        self, survey_sudo: Any, answer_sudo: Any, post: dict[str, Any]
    ) -> Any:
        if answer_sudo.is_session_answer:
            return survey_sudo.session_question_id

        next_page_or_question = None
        if "next_skipped_page" in post:
            next_page_or_question = answer_sudo._get_next_skipped_page_or_question()
        if not next_page_or_question:
            next_page_or_question = survey_sudo._get_next_page_or_question(
                answer_sudo,
                answer_sudo.last_displayed_page_id.id
                if answer_sudo.last_displayed_page_id
                else 0,
            )
            if not next_page_or_question:
                next_page_or_question = answer_sudo._get_next_skipped_page_or_question()
        return next_page_or_question

    def _get_last_page_triggering_answers(
        self,
        survey_sudo: Any,
        answer_sudo: Any,
        next_page_or_question: Any,
        triggered_questions_by_answer: Any,
    ) -> list[int]:
        pages_or_questions = survey_sudo._get_pages_or_questions(answer_sudo)
        following_questions = pages_or_questions.filtered(
            lambda pq: pq.sequence > next_page_or_question.sequence
        )
        next_page_suggested_answers = next_page_or_question.suggested_answer_ids
        if survey_sudo.questions_layout_effective == "page_per_section":
            following_questions = following_questions.question_ids
            next_page_suggested_answers = (
                next_page_or_question.question_ids.suggested_answer_ids
            )
        return [
            answer.id
            for answer in triggered_questions_by_answer
            if answer in next_page_suggested_answers
            and any(
                q in following_questions for q in triggered_questions_by_answer[answer]
            )
        ]

    def _prepare_question_html(
        self, survey_sudo: Any, answer_sudo: Any, **post: Any
    ) -> dict[str, Any]:
        survey_data = self._prepare_survey_data(survey_sudo, answer_sudo, **post)

        IrQweb = request.env["ir.qweb"].with_context(
            lang=self.env["res.lang"]._get_data(id=answer_sudo.lang_id.id).code
            or self._get_lang_with_fallback(answer_sudo.sudo(False)).code
        )
        if answer_sudo.state == "done":
            survey_content = IrQweb._render("survey.survey_fill_form_done", survey_data)
        else:
            survey_content = IrQweb._render(
                "survey.survey_fill_form_in_progress", survey_data
            )

        survey_progress = False
        if (
            answer_sudo.state == "in_progress"
            and not survey_data.get("question", request.env["survey.question"]).is_page
        ):
            if survey_sudo.questions_layout_effective == "page_per_section":
                page_ids = survey_sudo.page_ids.ids
                survey_progress = IrQweb._render(
                    "survey.survey_progression",
                    {
                        "survey": survey_sudo,
                        "page_ids": page_ids,
                        "page_number": page_ids.index(survey_data["page"].id)
                        + (1 if survey_sudo.progression_mode == "number" else 0),
                    },
                )
            elif survey_sudo.questions_layout_effective == "page_per_question":
                page_ids = (
                    answer_sudo.predefined_question_ids.ids
                    if not answer_sudo.is_session_answer
                    and survey_sudo.questions_selection == "random"
                    else survey_sudo.question_ids.ids
                )
                survey_progress = IrQweb._render(
                    "survey.survey_progression",
                    {
                        "survey": survey_sudo,
                        "page_ids": page_ids,
                        "page_number": page_ids.index(survey_data["question"].id),
                    },
                )

        background_image_url = survey_sudo.background_image_url
        if "question" in survey_data:
            background_image_url = survey_data["question"].background_image_url
        elif "page" in survey_data:
            background_image_url = survey_data["page"].background_image_url

        return {
            "has_skipped_questions": any(answer_sudo._get_skipped_questions()),
            "survey_content": survey_content,
            "survey_progress": survey_progress,
            "survey_navigation": IrQweb._render(
                "survey.survey_navigation", survey_data
            ),
            "background_image_url": background_image_url,
        }

    def _apply_url_prefill(
        self, survey_sudo: Any, answer_sudo: Any, post: dict[str, Any]
    ) -> None:
        question_ids = {q.id: q for q in survey_sudo.question_ids}
        existing_question_ids = set(
            answer_sudo.user_input_line_ids.mapped("question_id").ids
        )

        prefills = {}
        for key, value in post.items():
            if not value:
                continue
            question_id = None
            if key.startswith("prefill_"):
                try:
                    question_id = int(key.removeprefix("prefill_"))
                except ValueError:
                    continue
            elif key.startswith("Q") and key[1:].isdigit():
                question_id = int(key[1:])
            if (
                question_id
                and question_id in question_ids
                and question_id not in existing_question_ids
            ):
                prefills[question_id] = value

        for question_id, raw_value in prefills.items():
            question = question_ids[question_id]
            try:
                # A failure here can be a DB-level one (an out-of-int4 scale value
                # reaches PostgreSQL, not Python), which aborts the transaction. Without
                # a savepoint the swallow below leaves every later query in this request
                # dying on InFailedSqlTransaction.
                with request.env.cr.savepoint():
                    if question.question_type in ("simple_choice", "dropdown"):
                        answer = self._resolve_prefill_choice(question, raw_value)
                        if answer:
                            answer_sudo._save_lines(question, answer)
                    elif question.question_type == "multiple_choice":
                        answers = []
                        for part in raw_value.split(","):
                            ans = self._resolve_prefill_choice(question, part.strip())
                            if ans:
                                answers.append(ans)
                        if answers:
                            answer_sudo._save_lines(question, answers)
                    elif question.question_type in (
                        "char_box",
                        "text_box",
                        "numerical_box",
                        "date",
                        "datetime",
                        "scale",
                        "nps",
                        "slider",
                        "rating",
                    ):
                        answer_sudo._save_lines(question, raw_value)
            except Exception:
                _logger.debug(
                    "Skipping invalid prefill value for question %s",
                    question.id,
                    exc_info=True,
                )
                continue

    @staticmethod
    def _resolve_prefill_choice(question: Any, raw_value: str) -> int | None:
        try:
            answer_id = int(raw_value)
            if question.suggested_answer_ids.filtered(lambda a: a.id == answer_id):
                return answer_id
        except ValueError:
            pass
        for answer in question.suggested_answer_ids:
            if (answer.value or "").strip().lower() == raw_value.strip().lower():
                return answer.id
        return None

    @http.route(
        "/survey/<string:survey_token>/<string:answer_token>",
        type="http",
        auth="public",
        website=True,
    )
    def survey_display_page(
        self, survey_token: str, answer_token: str, **post: Any
    ) -> Response:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=True
        )
        if access_data["validity_code"] is not True:
            return self._redirect_with_error(access_data, access_data["validity_code"])

        answer_sudo = access_data["answer_sudo"]
        if answer_sudo.state != "done" and answer_sudo.survey_time_limit_reached:
            answer_sudo._mark_done()

        return request.render(
            "survey.survey_page_fill",
            self._prepare_survey_data(access_data["survey_sudo"], answer_sudo, **post),
        )

    @http.route(
        "/survey/<string:survey_token>/get_background_image",
        type="http",
        auth="public",
        website=True,
        sitemap=False,
    )
    def survey_get_background(self, survey_token: str) -> Response:
        survey_sudo, _dummy = self._fetch_from_access_token(survey_token, False)
        if not survey_sudo or not (
            survey_sudo.active
            or survey_sudo.with_user(request.env.user).has_access("read")
        ):
            raise werkzeug.exceptions.NotFound
        return (
            request.env["ir.binary"]
            ._get_stream_image_from_record(survey_sudo, "background_image")
            .get_response()
        )

    @http.route(
        "/survey/<string:survey_token>/<int:section_id>/get_background_image",
        type="http",
        auth="public",
        website=True,
        sitemap=False,
    )
    def survey_section_get_background(
        self, survey_token: str, section_id: int
    ) -> Response:
        survey_sudo, _dummy = self._fetch_from_access_token(survey_token, False)

        if not survey_sudo or not (
            survey_sudo.active
            or survey_sudo.with_user(request.env.user).has_access("read")
        ):
            raise werkzeug.exceptions.NotFound

        section = survey_sudo.page_ids.filtered(lambda q: q.id == section_id)
        if not section:
            raise werkzeug.exceptions.Forbidden

        return (
            request.env["ir.binary"]
            ._get_stream_image_from_record(section, "background_image")
            .get_response()
        )

    @http.route(
        "/survey/get_question_image/<string:survey_token>/<string:answer_token>/<int:question_id>/<int:suggested_answer_id>",
        type="http",
        auth="public",
        website=True,
        sitemap=False,
    )
    def survey_get_question_image(
        self,
        survey_token: str,
        answer_token: str,
        question_id: int,
        suggested_answer_id: int,
    ) -> Response:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=True
        )
        if access_data["validity_code"] is not True:
            return werkzeug.exceptions.Forbidden()

        survey_sudo = access_data["survey_sudo"]

        suggested_answer = False
        if int(question_id) in survey_sudo.question_ids.ids:
            suggested_answer = (
                request.env["survey.question.answer"]
                .sudo()
                .search(
                    [
                        ("id", "=", int(suggested_answer_id)),
                        ("question_id", "=", int(question_id)),
                        ("question_id.survey_id", "=", survey_sudo.id),
                    ]
                )
            )

        if not suggested_answer:
            return werkzeug.exceptions.NotFound()

        return (
            request.env["ir.binary"]
            ._get_stream_image_from_record(suggested_answer, "value_image")
            .get_response()
        )

    @http.route(
        "/survey/save_later/<string:survey_token>/<string:answer_token>",
        type="jsonrpc",
        auth="public",
        website=True,
    )
    def survey_save_later(
        self, survey_token: str, answer_token: str, **post: Any
    ) -> dict[str, Any]:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=True
        )
        if access_data["validity_code"] is not True:
            return {"error": access_data["validity_code"]}

        answer_sudo = access_data["answer_sudo"]

        if not answer_sudo.email:
            return {"error": "no_email"}

        if not answer_sudo._consume_save_later_allowance():
            return {"error": "too_many_requests"}

        template = self.env.ref("survey.mail_template_survey_save_later")
        template.sudo().send_mail(
            answer_sudo.id,
            email_values={"email_to": answer_sudo.email},
            force_send=True,
        )

        return {"success": True, "email": answer_sudo.email}

    @http.route(
        "/survey/upload/<string:survey_token>/<string:answer_token>",
        type="http",
        auth="public",
        methods=["POST"],
        csrf=True,
        website=True,
    )
    def survey_upload(
        self, survey_token: str, answer_token: str, **post: Any
    ) -> Response:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=True
        )
        if access_data["validity_code"] is not True:
            return request.make_json_response(
                {"error": access_data["validity_code"]}, status=403
            )
        answer_sudo = access_data["answer_sudo"]
        survey_sudo = access_data["survey_sudo"]

        try:
            question = survey_sudo.question_ids.browse(int(post.get("question_id")))
        except ValueError, TypeError:
            return request.make_json_response({"error": "invalid_question"}, status=400)
        if question not in survey_sudo.question_ids or (
            question.question_type != "file_upload"
        ):
            return request.make_json_response({"error": "invalid_question"}, status=400)

        upload = request.httprequest.files.get("file")
        if not upload:
            return request.make_json_response({"error": "no_file"}, status=400)

        content = upload.read(question.file_upload_max_size * 1024 * 1024 + 1)
        if len(content) > question.file_upload_max_size * 1024 * 1024:
            return request.make_json_response({"error": "too_large"}, status=413)

        # This route is auth="public" and creates a record per POST. Without a ceiling,
        # a respondent holding one valid answer token can fill the filestore one upload
        # at a time; the previous file for the same question is also released here, so
        # re-uploading does not accumulate.
        Attachment = request.env["ir.attachment"].sudo()
        own_domain = [
            ("res_model", "=", answer_sudo._name),
            ("res_id", "=", answer_sudo.id),
        ]
        if Attachment.search_count(own_domain) >= self.MAX_UPLOADS_PER_ANSWER:
            return request.make_json_response({"error": "too_many_files"}, status=429)

        previous = answer_sudo.user_input_line_ids.filtered(
            lambda line: line.question_id == question and line.value_char_box
        )
        superseded = Attachment.search(
            [
                *own_domain,
                (
                    "id",
                    "in",
                    [
                        int(line.value_char_box)
                        for line in previous
                        if line.value_char_box.isdigit()
                    ],
                ),
            ]
        )

        attachment = Attachment.create(
            {
                "name": upload.filename,
                "raw": content,
                "res_model": answer_sudo._name,
                "res_id": answer_sudo.id,
            }
        )
        errors = question._check_answer(attachment.id)
        if errors:
            attachment.unlink()
            return request.make_json_response(
                {"error": "rejected", "message": errors[question.id]}, status=400
            )
        superseded.unlink()
        return request.make_json_response(
            {"attachment_id": attachment.id, "name": attachment.name}
        )

    @http.route(
        "/survey/begin/<string:survey_token>/<string:answer_token>",
        type="jsonrpc",
        auth="public",
        website=True,
    )
    def survey_begin(
        self, survey_token: str, answer_token: str, **post: Any
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=True
        )
        if access_data["validity_code"] is not True:
            return {}, {"error": access_data["validity_code"]}
        survey_sudo, answer_sudo = (
            access_data["survey_sudo"],
            access_data["answer_sudo"],
        )

        if answer_sudo.state != "new":
            return {}, {"error": _("The survey has already started.")}

        if "lang_code" in post:
            lang = request.env["res.lang"]._lang_get(post["lang_code"])
            if lang:
                answer_sudo.lang_id = lang
        answer_sudo._mark_in_progress()

        self._apply_url_prefill(survey_sudo, answer_sudo, post)

        return {}, self._prepare_question_html(survey_sudo, answer_sudo, **post)

    @http.route(
        "/survey/next_question/<string:survey_token>/<string:answer_token>",
        type="jsonrpc",
        auth="public",
        website=True,
    )
    def survey_next_question(
        self, survey_token: str, answer_token: str, **post: Any
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=True
        )
        if access_data["validity_code"] is not True:
            return {}, {"error": access_data["validity_code"]}
        survey_sudo, answer_sudo = (
            access_data["survey_sudo"],
            access_data["answer_sudo"],
        )

        if answer_sudo.state == "new" and answer_sudo.is_session_answer:
            answer_sudo._mark_in_progress()

        return {}, self._prepare_question_html(survey_sudo, answer_sudo, **post)

    def _check_time_limit_exceeded(self, survey_sudo: Any, answer_sudo: Any) -> bool:
        if not (
            answer_sudo.survey_time_limit_reached
            or answer_sudo.question_time_limit_reached
        ):
            return False
        if answer_sudo.question_time_limit_reached:
            time_limit = survey_sudo.session_question_start_time + relativedelta(
                seconds=survey_sudo.session_question_id.time_limit
            )
            time_limit += timedelta(seconds=3)
        else:
            time_limit = answer_sudo.start_datetime + timedelta(
                minutes=survey_sudo.time_limit
            )
            time_limit += timedelta(seconds=10)
        return fields.Datetime.now() > time_limit

    def _determine_next_page_after_submit(
        self,
        survey_sudo: Any,
        answer_sudo: Any,
        page_or_question_id: int,
        correct_answers: dict[str, Any],
        **post: Any,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        if "previous_page_id" in post:
            answer_sudo.last_displayed_page_id = post["previous_page_id"]
            return correct_answers, self._prepare_question_html(
                survey_sudo, answer_sudo, **post
            )

        if "next_skipped_page_or_question" in post:
            answer_sudo.last_displayed_page_id = page_or_question_id
            return correct_answers, self._prepare_question_html(
                survey_sudo, answer_sudo, next_skipped_page=True
            )

        if not answer_sudo.is_session_answer:
            skip_result = self._check_skip_actions(
                survey_sudo, answer_sudo, page_or_question_id, correct_answers
            )
            if skip_result is not None:
                return skip_result

            if answer_sudo.survey_first_submitted:
                next_page = request.env["survey.question"]
            else:
                next_page = survey_sudo._get_next_page_or_question(
                    answer_sudo, page_or_question_id
                )
            if not next_page:
                if (
                    survey_sudo.users_can_go_back
                    and answer_sudo.user_input_line_ids.filtered(
                        lambda a: a.skipped and a.question_id.constr_mandatory
                    )
                ):
                    answer_sudo.write(
                        {
                            "last_displayed_page_id": page_or_question_id,
                            "survey_first_submitted": True,
                        }
                    )
                    return correct_answers, self._prepare_question_html(
                        survey_sudo, answer_sudo, next_skipped_page=True
                    )
                else:
                    answer_sudo._mark_done()

        answer_sudo.last_displayed_page_id = page_or_question_id
        return correct_answers, self._prepare_question_html(survey_sudo, answer_sudo)

    @http.route(
        "/survey/submit/<string:survey_token>/<string:answer_token>",
        type="jsonrpc",
        auth="public",
        website=True,
    )
    def survey_submit(
        self, survey_token: str, answer_token: str, **post: Any
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=True
        )
        if access_data["validity_code"] is not True:
            return {}, {"error": access_data["validity_code"]}
        survey_sudo, answer_sudo = (
            access_data["survey_sudo"],
            access_data["answer_sudo"],
        )

        # Taken before the state is read, so a second request carrying the same token
        # waits here rather than racing this one to _mark_done().
        answer_sudo._lock()
        answer_sudo.invalidate_recordset(["state"])
        if answer_sudo.state == "done":
            return {}, {"error": "unauthorized"}

        questions, page_or_question_id = survey_sudo._get_survey_questions(
            answer=answer_sudo,
            page_id=post.get("page_id"),
            question_id=post.get("question_id"),
        )

        if not answer_sudo.test_entry and not survey_sudo._has_attempts_left(
            answer_sudo.partner_id, answer_sudo.email, answer_sudo.invite_token
        ):
            return {}, {"error": "unauthorized"}

        if self._check_time_limit_exceeded(survey_sudo, answer_sudo):
            return {}, {"error": "unauthorized"}

        errors = self._validate_and_save_page(survey_sudo, answer_sudo, questions, post)

        if errors and not (
            answer_sudo.survey_time_limit_reached
            or answer_sudo.question_time_limit_reached
        ):
            return {}, {"error": "validation", "fields": errors}

        if not answer_sudo.is_session_answer:
            answer_sudo._clear_inactive_conditional_answers()

        answer_sudo._evaluate_calculated_fields()

        answer_sudo._fire_webhook("page_submitted")

        correct_answers = {}
        if survey_sudo.scoring_type == "scoring_with_answers_after_page":
            scorable_questions = (
                questions - answer_sudo._get_inactive_conditional_questions()
            ).filtered("is_scored_question")
            correct_answers = scorable_questions._get_correct_answers()

        if (
            answer_sudo.survey_time_limit_reached
            or survey_sudo.questions_layout_effective == "one_page"
        ):
            answer_sudo._mark_done()
            return correct_answers, self._prepare_question_html(
                survey_sudo, answer_sudo
            )

        return self._determine_next_page_after_submit(
            survey_sudo, answer_sudo, page_or_question_id, correct_answers, **post
        )

    def _validate_and_save_page(
        self, survey_sudo: Any, answer_sudo: Any, questions: Any, post: dict[str, Any]
    ) -> dict[int, str]:
        gating_questions = (
            survey_sudo.question_and_page_ids.triggering_answer_ids.question_id
            | survey_sudo.question_and_page_ids.triggering_question_id
        )
        inactive_questions = (
            request.env["survey.question"]
            if answer_sudo.is_session_answer
            else answer_sudo._get_inactive_conditional_questions()
        )
        errors = {}
        for question in questions:
            if question in inactive_questions:
                continue
            answer, comment = self._extract_comment_from_answers(
                question, post.get(str(question.id))
            )
            errors.update(question._check_answer(answer, comment))
            if errors.get(question.id):
                continue
            if quota_error := self._check_answer_quota(survey_sudo, question, answer):
                errors[question.id] = quota_error
                continue
            answer_sudo._save_lines(
                question,
                answer,
                comment,
                overwrite_existing=survey_sudo.users_can_go_back
                or question.save_as_nickname
                or question.save_as_email,
            )
            if question in gating_questions and not answer_sudo.is_session_answer:
                inactive_questions = answer_sudo._get_inactive_conditional_questions()
        return errors

    def _check_answer_quota(
        self, survey_sudo: Any, question: Any, answer: Any
    ) -> str | None:
        if not survey_sudo.quota_ids or not answer:
            return None
        if question.question_type not in (
            "simple_choice",
            "dropdown",
            "multiple_choice",
        ):
            return None
        answer_ids = (
            [int(a) for a in answer] if isinstance(answer, list) else [int(answer)]
        )
        if survey_sudo.quota_ids._check_quota(answer_ids):
            return _("One or more selected answers have reached their response quota.")
        return None

    _SKIP_ACTION_PRECEDENCE = {"end_survey": 0, "redirect": 1, "skip_to": 2}

    def _check_skip_actions(
        self,
        survey_sudo: Any,
        answer_sudo: Any,
        page_or_question_id: int,
        correct_answers: dict[str, Any],
    ) -> tuple[dict[str, Any], dict[str, Any]] | None:
        Question = survey_sudo.env["survey.question"]
        if survey_sudo.questions_layout_effective == "page_per_question":
            submitted_questions = Question.browse(page_or_question_id)
        else:
            page = Question.browse(page_or_question_id)
            submitted_questions = page.question_ids if page.is_page else page

        selected = answer_sudo.user_input_line_ids.filtered(
            lambda ln: ln.question_id in submitted_questions and ln.suggested_answer_id
        ).suggested_answer_id
        # A multiple-choice question can carry several answers with different skip
        # actions. Taking whichever line the recordset happened to yield first made the
        # winner an artifact of insertion order; the most terminal action wins now, so
        # picking "end the survey" alongside "jump to Q7" ends the survey.
        for answer in selected.sorted(
            lambda a: (
                self._SKIP_ACTION_PRECEDENCE.get(a.skip_action, 99),
                a.sequence,
                a.id,
            )
        ):
            if answer.skip_action == "end_survey":
                answer_sudo._mark_done()
                return correct_answers, self._prepare_question_html(
                    survey_sudo, answer_sudo
                )
            elif answer.skip_action == "redirect" and answer.skip_redirect_url:
                answer_sudo._mark_done()
                return correct_answers, {
                    "redirect_url": answer.skip_redirect_url,
                }
            elif answer.skip_action == "skip_to" and answer.skip_target_id:
                target = answer.skip_target_id
                if target.id in survey_sudo.question_and_page_ids.ids:
                    before_target = survey_sudo._get_next_page_or_question(
                        answer_sudo, target.id, go_back=True
                    )
                    answer_sudo.last_displayed_page_id = before_target or False
                    return correct_answers, self._prepare_question_html(
                        survey_sudo, answer_sudo
                    )
        return None

    def _extract_comment_from_answers(
        self, question: Any, answers: Any
    ) -> tuple[Any, str | None]:
        comment = None
        answers_no_comment = []
        if not question._is_well_shaped_answer(answers):
            # Hand the payload on unchanged; _check_answer refuses it with a message.
            # Reaching into it here -- `"comment" in answers` on a number -- raised out
            # of the request instead.
            return answers, None
        if not question._is_unanswered(answers):
            if question.question_type in (
                "matrix",
                "likert",
                "ranking",
                "constant_sum",
            ):
                if "comment" in answers:
                    comment = answers["comment"].strip()
                    answers.pop("comment")
                answers_no_comment = answers
            else:
                if not isinstance(answers, list):
                    answers = [answers]
                for answer in answers:
                    if isinstance(answer, dict) and "comment" in answer:
                        comment = answer["comment"].strip()
                    else:
                        answers_no_comment.append(answer)
                if len(answers_no_comment) == 1:
                    answers_no_comment = answers_no_comment[0]
        return answers_no_comment, comment

    @http.route(
        "/survey/print/<string:survey_token>",
        type="http",
        auth="public",
        website=True,
        sitemap=False,
    )
    def survey_print(
        self,
        survey_token: str,
        review: bool = False,
        answer_token: str | None = None,
        **post: Any,
    ) -> Response:
        access_data = self._get_access_data(
            survey_token, answer_token, ensure_token=False, check_partner=False
        )
        if access_data["validity_code"] is not True and (
            not access_data["has_survey_access"]
            or access_data["validity_code"]
            not in ["token_required", "survey_closed", "survey_void", "answer_deadline"]
        ):
            return self._redirect_with_error(access_data, access_data["validity_code"])

        survey_sudo, answer_sudo = (
            access_data["survey_sudo"],
            access_data["answer_sudo"],
        )
        return request.render(
            "survey.survey_page_print",
            {
                "is_html_empty": is_html_empty,
                "review": review,
                "survey": survey_sudo,
                "answer": answer_sudo
                if survey_sudo.scoring_type != "scoring_without_answers"
                else answer_sudo.browse(),
                "questions_to_display": answer_sudo._get_print_questions(),
                "scoring_display_correction": survey_sudo.scoring_type
                in ["scoring_with_answers", "scoring_with_answers_after_page"]
                and answer_sudo,
                "format_datetime": lambda dt: format_datetime(
                    request.env, dt, dt_format=False
                ),
                "format_date": lambda date: format_date(request.env, date),
                "graph_data": json.dumps(
                    answer_sudo._prepare_answer_statistics()[answer_sudo]
                )
                if answer_sudo
                and survey_sudo.scoring_type
                in ["scoring_with_answers", "scoring_with_answers_after_page"]
                else False,
            },
        )

    @http.route(
        '/survey/<model("survey.survey"):survey>/certification_preview',
        type="http",
        auth="user",
        website=True,
    )
    def show_certification_pdf(self, survey: Any, **kwargs: Any) -> Response:
        preview_url = f"/survey/{survey.id}/get_certification_preview"
        return request.render(
            "survey.certification_preview",
            {
                "preview_url": preview_url,
                "page_title": survey.title,
            },
        )

    @http.route(
        ['/survey/<model("survey.survey"):survey>/get_certification_preview'],
        type="http",
        auth="user",
        methods=["GET"],
        website=True,
    )
    def survey_get_certification_preview(self, survey: Any, **kwargs: Any) -> Response:
        if not request.env.user.has_group("survey.group_survey_user"):
            raise werkzeug.exceptions.Forbidden

        fake_user_input = survey._create_answer(user=request.env.user, test_entry=True)
        try:
            response = self._generate_report(fake_user_input, download=False)
        finally:
            fake_user_input.sudo().unlink()
        return response

    @http.route(
        ["/survey/<int:survey_id>/get_certification"],
        type="http",
        auth="user",
        methods=["GET"],
        website=True,
    )
    def survey_get_certification(self, survey_id: int, **kwargs: Any) -> Response:
        survey = (
            request.env["survey.survey"]
            .sudo()
            .search([("id", "=", survey_id), ("certification", "=", True)])
        )

        if not survey:
            return request.redirect("/")

        succeeded_attempt = (
            request.env["survey.user_input"]
            .sudo()
            .search(
                [
                    ("partner_id", "=", request.env.user.partner_id.id),
                    ("survey_id", "=", survey_id),
                    ("scoring_success", "=", True),
                    # _mark_done already refuses to mail a certificate for a test
                    # entry; downloading one had no such rule, so a test run that
                    # scored well produced a real certificate.
                    ("test_entry", "=", False),
                ],
                # The best passing attempt, not whichever the default order returned.
                order="scoring_percentage desc, id desc",
                limit=1,
            )
        )

        if not succeeded_attempt:
            return request.redirect("/")

        return self._generate_report(succeeded_attempt, download=True)

    @http.route(
        '/survey/results/<model("survey.survey"):survey>',
        type="http",
        auth="user",
        website=True,
    )
    def survey_report(
        self, survey: Any, answer_token: str | None = None, **post: Any
    ) -> Response:
        user_input_lines, search_filters = self._extract_filters_data(survey, post)
        survey_data = survey._prepare_survey_statistics(user_input_lines)
        question_and_page_data = (
            survey.question_and_page_ids._prepare_question_statistics(user_input_lines)
        )

        template_values = {
            "survey": survey,
            "question_and_page_data": question_and_page_data,
            "survey_data": survey_data,
            "search_filters": search_filters,
            "search_finished": post.get("finished") == "true",
            "search_failed": post.get("failed") == "true",
            "search_passed": post.get("passed") == "true",
        }

        if survey.session_show_leaderboard:
            template_values["leaderboard"] = survey._prepare_leaderboard_values()

        return request.render("survey.survey_page_statistics", template_values)

    @http.route(
        '/survey/results/<model("survey.survey"):survey>/cross_tabulation',
        type="jsonrpc",
        auth="user",
    )
    def survey_cross_tabulation(
        self, survey: Any, question_row_id: int, question_col_id: int
    ) -> dict[str, Any]:
        return survey._prepare_cross_tabulation(question_row_id, question_col_id)

    def _generate_report(self, user_input: Any, download: bool = True) -> Response:
        report = (
            request.env["ir.actions.report"]
            .sudo()
            ._render_qweb_pdf(
                "survey.certification_report",
                [user_input.id],
                data={"report_type": "pdf"},
            )[0]
        )

        report_content_disposition = content_disposition("Certification.pdf")
        if not download:
            content_split = report_content_disposition.split(";")
            content_split[0] = "inline"
            report_content_disposition = ";".join(content_split)

        return request.make_response(
            report,
            headers=[
                ("Content-Type", "application/pdf"),
                ("Content-Length", len(report)),
                ("Content-Disposition", report_content_disposition),
            ],
        )

    def _get_results_page_user_input_domain(self, survey: Any, **post: Any) -> Domain:
        user_input_domains = []
        if post.get("finished"):
            user_input_domains.append(Domain("state", "=", "done"))
        else:
            user_input_domains.append(Domain("state", "!=", "new"))
        if post.get("failed"):
            user_input_domains.append(Domain("scoring_success", "=", False))
        elif post.get("passed"):
            user_input_domains.append(Domain("scoring_success", "=", True))

        if post.get("date_from"):
            user_input_domains.append(Domain("end_datetime", ">=", post["date_from"]))
        if post.get("date_to"):
            user_input_domains.append(Domain("end_datetime", "<=", post["date_to"]))

        if post.get("score_min"):
            with contextlib.suppress(ValueError):
                user_input_domains.append(
                    Domain("scoring_percentage", ">=", float(post["score_min"]))
                )
        if post.get("score_max"):
            with contextlib.suppress(ValueError):
                user_input_domains.append(
                    Domain("scoring_percentage", "<=", float(post["score_max"]))
                )

        if post.get("quality_min"):
            with contextlib.suppress(ValueError):
                user_input_domains.append(
                    Domain("quality_score", ">=", int(post["quality_min"]))
                )

        user_input_domains.extend(
            (Domain("test_entry", "=", False), Domain("survey_id", "=", survey.id))
        )
        return Domain.AND(user_input_domains)

    def _extract_filters_data(
        self, survey: Any, post: dict[str, Any]
    ) -> tuple[Any, list[dict[str, Any]]]:
        user_input_line_subdomains = []
        search_filters = []

        answer_by_column, user_input_lines_ids = self._get_filters_from_post(post)

        if answer_by_column:
            answer_ids, row_ids = [], []
            for answer_column_id, answer_row_ids in answer_by_column.items():
                answer_ids.append(answer_column_id)
                row_ids += answer_row_ids

            answers_and_rows = request.env["survey.question.answer"].browse(
                answer_ids + row_ids
            )
            answers = answers_and_rows.filtered(lambda a: not a.matrix_question_id)

            for answer in answers:
                if not answer_by_column[answer.id]:
                    user_input_line_subdomains.append(
                        answer._get_answer_matching_domain()
                    )
                    search_filters.append(self._prepare_search_filter_answer(answer))
                else:
                    for row_id in answer_by_column[answer.id]:
                        row = answers_and_rows.filtered(
                            lambda answer_or_row, rid=row_id: answer_or_row.id == rid
                        )
                        user_input_line_subdomains.append(
                            answer._get_answer_matching_domain(row_id)
                        )
                        search_filters.append(
                            self._prepare_search_filter_answer(answer, row)
                        )

        if user_input_lines_ids:
            user_input_lines = request.env["survey.user_input.line"].browse(
                user_input_lines_ids
            )
            for input_line in user_input_lines:
                user_input_line_subdomains.append(
                    input_line._get_answer_matching_domain()
                )
                search_filters.append(
                    self._prepare_search_filter_input_line(input_line)
                )

        user_input_domain = self._get_results_page_user_input_domain(survey, **post)

        if user_input_line_subdomains:
            all_required_lines_domains = [
                [
                    (
                        "user_input_line_ids",
                        "in",
                        request.env["survey.user_input.line"].sudo()._search(subdomain),
                    )
                ]
                for subdomain in user_input_line_subdomains
            ]
            user_input_domain = Domain.AND(
                [user_input_domain, *all_required_lines_domains]
            )

        user_inputs_query = (
            request.env["survey.user_input"].sudo()._search(user_input_domain)
        )
        user_input_lines = request.env["survey.user_input.line"].search(
            [("user_input_id", "in", user_inputs_query)]
        )

        return user_input_lines, search_filters

    def _get_filters_from_post(
        self, post: dict[str, Any]
    ) -> tuple[defaultdict[int, list[int]], list[int]]:
        answer_by_column = defaultdict(list)
        user_input_lines_ids = []

        for data in post.get("filters", "").split("|"):
            if not data:
                break
            parts = data.split(",")
            if len(parts) != 3:
                continue
            model_short_key, row_id, answer_id = parts
            try:
                row_id, answer_id = int(row_id), int(answer_id)
            except ValueError, TypeError:
                continue
            if model_short_key == "A":
                if row_id:
                    answer_by_column[answer_id].append(row_id)
                else:
                    answer_by_column[answer_id] = []
            elif model_short_key == "L" and not row_id:
                user_input_lines_ids.append(answer_id)

        return answer_by_column, user_input_lines_ids

    def _prepare_search_filter_answer(
        self, answer: Any, row: Any = False
    ) -> dict[str, Any]:
        return {
            "question_id": answer.question_id.id,
            "question": answer.question_id.title,
            "row_id": row.id if row else 0,
            "answer": f"{row.value} : {answer.value}" if row else answer.value,
            "model_short_key": "A",
            "record_id": answer.id,
        }

    def _prepare_search_filter_input_line(self, user_input_line: Any) -> dict[str, Any]:
        return {
            "question_id": user_input_line.question_id.id,
            "question": user_input_line.question_id.title,
            "row_id": 0,
            "answer": user_input_line._get_answer_value(),
            "model_short_key": "L",
            "record_id": user_input_line.id,
        }

    def _get_lang_with_fallback(self, user_input: Any) -> Any:
        user_input.ensure_one()
        user_input_sudo = user_input.sudo()
        if user_input_sudo.lang_id:
            return user_input_sudo.lang_id.sudo(False)
        lang_code = (
            self.env.context.get("lang") or self.env["ir.http"]._get_default_lang().code
        )
        ResLang = self.env["res.lang"]
        supported_lang_codes = user_input_sudo.survey_id._get_supported_lang_codes()
        supported_lang_codes_set = set(supported_lang_codes)
        if lang_code in supported_lang_codes_set:
            return ResLang._lang_get(lang_code)
        return ResLang._lang_get(
            next(
                (
                    lang.code
                    for lang in self.env["res.lang"]._get_frontend().values()
                    if lang["code"] in supported_lang_codes_set
                ),
                supported_lang_codes[0],
            )
        )

    @http.route(
        '/survey/results/<model("survey.survey"):survey>/export/csv',
        type="http",
        auth="user",
    )
    def survey_export_csv(self, survey: Any, **post: Any) -> Response:
        header, rows = self._build_export_data(survey)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)

        filename = f"{survey.title} - Responses.csv"
        return request.make_response(
            output.getvalue(),
            headers=[
                ("Content-Type", "text/csv;charset=utf-8"),
                ("Content-Disposition", content_disposition(filename)),
            ],
        )

    @http.route(
        '/survey/results/<model("survey.survey"):survey>/export/xlsx',
        type="http",
        auth="user",
    )
    def survey_export_xlsx(self, survey: Any, **post: Any) -> Response:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header, rows = self._build_export_data(survey)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Responses"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="714B67", end_color="714B67", fill_type="solid"
        )
        for col_idx, col_name in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx in range(1, len(header) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, min(len(rows) + 2, 50))
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 2, 40
            )

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)

        filename = f"{survey.title} - Responses.xlsx"
        return request.make_response(
            output.getvalue(),
            headers=[
                (
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                ("Content-Disposition", content_disposition(filename)),
            ],
        )

    def _build_export_data(self, survey: Any) -> tuple[list[str], list[list]]:
        user_inputs = request.env["survey.user_input"].search(
            [
                ("survey_id", "=", survey.id),
                ("state", "=", "done"),
                ("test_entry", "=", False),
            ]
        )
        questions = survey.question_ids

        header = [
            "Respondent",
            "Email",
            "Start Date",
            "End Date",
            "Duration (min)",
            "Score (%)",
        ]
        for question in questions:
            if question.question_type in ("matrix", "likert"):
                header.extend(
                    f"{question.title} [{row.value}]" for row in question.matrix_row_ids
                )
            else:
                header.append(question.title)

        rows = []
        for user_input in user_inputs:
            row = [
                user_input.nickname
                or (user_input.partner_id.name if user_input.partner_id else ""),
                user_input.email or "",
                str(user_input.start_datetime or ""),
                str(user_input.end_datetime or ""),
                round(
                    (
                        user_input.end_datetime - user_input.start_datetime
                    ).total_seconds()
                    / 60,
                    1,
                )
                if user_input.start_datetime and user_input.end_datetime
                else "",
                round(user_input.scoring_percentage, 1)
                if survey.scoring_type != "no_scoring"
                else "",
            ]

            lines_by_question = user_input.user_input_line_ids.grouped("question_id")
            for question in questions:
                q_lines = lines_by_question.get(
                    question, request.env["survey.user_input.line"]
                )
                if question.question_type in ("matrix", "likert"):
                    lines_by_row = q_lines.grouped("matrix_row_id")
                    for matrix_row in question.matrix_row_ids:
                        row_lines = lines_by_row.get(
                            matrix_row, request.env["survey.user_input.line"]
                        )
                        row.append(
                            ", ".join(
                                ln.suggested_answer_id.value
                                for ln in row_lines
                                if ln.suggested_answer_id
                            )
                        )
                elif question.question_type in (
                    "simple_choice",
                    "dropdown",
                    "multiple_choice",
                ):
                    row.append(
                        ", ".join(
                            ln.suggested_answer_id.value
                            for ln in q_lines
                            if ln.suggested_answer_id
                        )
                    )
                elif question.question_type == "text_box":
                    row.append(q_lines[0].value_text_box if q_lines else "")
                elif question.question_type == "char_box":
                    row.append(q_lines[0].value_char_box if q_lines else "")
                elif question.question_type == "numerical_box":
                    row.append(q_lines[0].value_numerical_box if q_lines else "")
                elif question.question_type in ("scale", "nps", "rating"):
                    row.append(q_lines[0].value_scale if q_lines else "")
                elif question.question_type == "slider":
                    row.append(q_lines[0].value_numerical_box if q_lines else "")
                elif question.question_type in ("ranking", "constant_sum"):
                    row.append(
                        ", ".join(
                            f"{ln.suggested_answer_id.value}: {ln.value_numerical_box}"
                            for ln in q_lines
                            if ln.suggested_answer_id and not ln.skipped
                        )
                    )
                elif question.question_type == "file_upload":
                    row.append(q_lines[0].value_char_box if q_lines else "")
                elif question.question_type == "date":
                    row.append(
                        str(q_lines[0].value_date)
                        if q_lines and q_lines[0].value_date
                        else ""
                    )
                elif question.question_type == "datetime":
                    row.append(
                        str(q_lines[0].value_datetime)
                        if q_lines and q_lines[0].value_datetime
                        else ""
                    )
                else:
                    row.append("")
            rows.append([self._sanitize_export_cell(cell) for cell in row])

        return header, rows

    _FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")

    @staticmethod
    def _sanitize_export_cell(value: Any) -> Any:
        if isinstance(value, str) and value.startswith(Survey._FORMULA_TRIGGERS):
            return f"'{value}"
        return value

    @http.route(
        '/survey/results/<model("survey.survey"):survey>/cross-tab',
        type="http",
        auth="user",
        website=True,
    )
    def survey_cross_tab(self, survey: Any, **post: Any) -> Response:
        choice_questions = survey.question_ids.filtered(
            lambda q: (
                q.question_type
                in (
                    "simple_choice",
                    "dropdown",
                    "multiple_choice",
                    "scale",
                    "nps",
                    "rating",
                )
            )
        )

        cross_tab_data = {}
        try:
            q_row = int(post.get("q_row") or 0)
            q_col = int(post.get("q_col") or 0)
        except ValueError, TypeError:
            q_row = q_col = 0
        if q_row and q_col and q_row != q_col:
            cross_tab_data = survey._prepare_cross_tabulation(q_row, q_col)

        return request.render(
            "survey.survey_page_cross_tab",
            {
                "survey": survey,
                "choice_questions": choice_questions,
                "cross_tab_data": cross_tab_data,
                "selected_row": q_row,
                "selected_col": q_col,
            },
        )

    @http.route(
        '/survey/results/<model("survey.survey"):survey>/segments',
        type="jsonrpc",
        auth="user",
    )
    def survey_segments(self, survey: Any, **post: Any) -> dict[str, Any]:
        request.env.cr.execute(
            """
            SELECT
                scoring_percentage,
                quality_score,
                EXTRACT(EPOCH FROM (end_datetime - start_datetime)) / 60.0 AS duration_min
            FROM survey_user_input
            WHERE survey_id = %s
              AND state = 'done'
              AND test_entry = FALSE
              AND end_datetime IS NOT NULL
              AND start_datetime IS NOT NULL
        """,
            [survey.id],
        )
        rows = request.env.cr.fetchall()
        if not rows:
            return {
                "score_bands": [],
                "quality_tiers": [],
                "duration_buckets": [],
                "total": 0,
            }

        bands = dict.fromkeys(("0-25%", "25-50%", "50-75%", "75-100%"), 0)
        for score, _quality, _dur in rows:
            if score < 25:
                bands["0-25%"] += 1
            elif score < 50:
                bands["25-50%"] += 1
            elif score < 75:
                bands["50-75%"] += 1
            else:
                bands["75-100%"] += 1

        tiers = dict.fromkeys(("low", "medium", "high"), 0)
        for _score, quality, _dur in rows:
            if quality <= 33:
                tiers["low"] += 1
            elif quality <= 66:
                tiers["medium"] += 1
            else:
                tiers["high"] += 1

        durations = sorted(d for _s, _q, d in rows if d and d > 0)
        if durations:
            q1 = durations[len(durations) // 4]
            median = durations[len(durations) // 2]
            q3 = durations[3 * len(durations) // 4]
            buckets = {
                _("< %(minutes).0f min", minutes=q1): len(
                    [d for d in durations if d < q1]
                ),
                _("%(low).0f-%(high).0f min", low=q1, high=median): len(
                    [d for d in durations if q1 <= d < median]
                ),
                _("%(low).0f-%(high).0f min", low=median, high=q3): len(
                    [d for d in durations if median <= d < q3]
                ),
                _("> %(minutes).0f min", minutes=q3): len(
                    [d for d in durations if d >= q3]
                ),
            }
        else:
            buckets = {}

        return {
            "score_bands": [{"label": k, "count": v} for k, v in bands.items()],
            "quality_tiers": [
                {"label": label, "count": tiers[key]}
                for key, label in (
                    ("low", _("Low (0-33)")),
                    ("medium", _("Medium (34-66)")),
                    ("high", _("High (67-100)")),
                )
            ],
            "duration_buckets": [{"label": k, "count": v} for k, v in buckets.items()],
            "total": len(rows),
        }

    @http.route(
        '/survey/results/<model("survey.survey"):survey>/compare',
        type="jsonrpc",
        auth="user",
    )
    def survey_compare(
        self,
        survey: Any,
        period_a_from: str = "",
        period_a_to: str = "",
        period_b_from: str = "",
        period_b_to: str = "",
        **post: Any,
    ) -> dict[str, Any]:
        def _get_lines_for_period(date_from, date_to):
            domain = [
                ("survey_id", "=", survey.id),
                ("state", "=", "done"),
                ("test_entry", "=", False),
            ]
            if date_from:
                domain.append(("end_datetime", ">=", date_from))
            if date_to:
                domain.append(("end_datetime", "<=", date_to))
            user_inputs = request.env["survey.user_input"].sudo().search(domain)
            return {
                "count": len(user_inputs),
                "avg_score": round(
                    sum(user_inputs.mapped("scoring_percentage"))
                    / (len(user_inputs) or 1),
                    1,
                ),
                "avg_quality": round(
                    sum(user_inputs.mapped("quality_score")) / (len(user_inputs) or 1),
                    1,
                ),
                "success_rate": round(
                    len(user_inputs.filtered("scoring_success"))
                    / (len(user_inputs) or 1)
                    * 100,
                    1,
                )
                if survey.scoring_type != "no_scoring"
                else None,
            }

        period_a = _get_lines_for_period(period_a_from, period_a_to)
        period_b = _get_lines_for_period(period_b_from, period_b_to)

        deltas = {}
        for key in ("count", "avg_score", "avg_quality", "success_rate"):
            a_val = period_a.get(key)
            b_val = period_b.get(key)
            if a_val is not None and b_val is not None:
                deltas[key] = round(b_val - a_val, 1)
            else:
                deltas[key] = None

        return {
            "period_a": period_a,
            "period_b": period_b,
            "deltas": deltas,
        }

    @http.route(
        '/survey/results/<model("survey.survey"):survey>/trends',
        type="jsonrpc",
        auth="user",
    )
    def survey_trends(
        self, survey: Any, granularity: str = "day", **post: Any
    ) -> dict[str, Any]:
        if granularity not in ("day", "week", "month"):
            granularity = "day"

        trunc = {"day": "day", "week": "week", "month": "month"}[granularity]

        request.env.cr.execute(
            """
            SELECT
                date_trunc(%s, end_datetime) AS period,
                COUNT(*) AS cnt,
                AVG(scoring_percentage) AS avg_score
            FROM survey_user_input
            WHERE survey_id = %s
              AND state = 'done'
              AND test_entry = FALSE
              AND end_datetime IS NOT NULL
            GROUP BY period
            ORDER BY period
        """,
            [trunc, survey.id],
        )

        results = request.env.cr.fetchall()
        has_scoring = survey.scoring_type != "no_scoring"

        date_fmt = {
            "day": "%Y-%m-%d",
            "week": "%Y-W%W",
            "month": "%Y-%m",
        }[granularity]

        return {
            "labels": [row[0].strftime(date_fmt) for row in results],
            "counts": [row[1] for row in results],
            "avg_scores": [round(row[2] or 0, 1) for row in results]
            if has_scoring
            else [],
            "granularity": granularity,
        }
